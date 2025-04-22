from flask import Flask, render_template, Response, jsonify, request, session, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import cv2
import face_recognition
import numpy as np
import pickle
import base64
import json
import os
from datetime import datetime, timedelta
import mysql.connector
from werkzeug.security import generate_password_hash, check_password_hash
from sklearn import neighbors
import time
from PIL import Image
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import tempfile

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here'  # Change this to a secure secret key

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

class User(UserMixin):
    def __init__(self, user_data):
        self.id = user_data['id']
        self.username = user_data['username']
        self.role = user_data['role']
        
    def get_id(self):
        return str(self.id)

@login_manager.user_loader
def load_user(user_id):
    print(f"Loading user with ID: {user_id}")
    conn = get_db_connection()
    if conn is None:
        print("Database connection failed in load_user")
        return None
    cursor = conn.cursor(dictionary=True)
    cursor.execute('SELECT * FROM users WHERE id = %s', (user_id,))
    user_data = cursor.fetchone()
    cursor.close()
    conn.close()
    if user_data:
        print(f"User loaded successfully: {user_data['username']}")
        return User(user_data)
    print(f"No user found with ID: {user_id}")
    return None

# Database configuration
with open('config.json', 'r', encoding='utf-8') as f:
    config = json.load(f)

db_username = base64.b64decode(config.get("db_username", "")).decode("utf-8")
db_password = base64.b64decode(config.get("db_password", "")).decode("utf-8")
db_host = config.get("db_host", "localhost")

# Initialize MySQL connection
def get_db_connection():
    try:
        conn = mysql.connector.connect(
            host=db_host,
            user=db_username,
            password=db_password,
            database="attendance_db"
        )
        return conn
    except mysql.connector.Error as err:
        return None

# Initialize face recognition model
def load_face_recognition_model():
    model_path = "trained_knn_model.clf"
    if os.path.exists(model_path):
        with open(model_path, 'rb') as f:
            return pickle.load(f)
    return None

# Routes
@app.route('/')
def index():
    print("Index route accessed")
    if not current_user.is_authenticated:
        print("User not authenticated, redirecting to login")
        return redirect(url_for('login'))
    print(f"User authenticated: {current_user.username}, role: {current_user.role}")
    return render_template('dashboard.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        print(f"Login attempt for user: {username}")
        
        conn = get_db_connection()
        if conn is None:
            print("Database connection failed during login")
            return render_template('login.html', error='Database connection failed. Please check your configuration.')
            
        cursor = conn.cursor(dictionary=True)
        cursor.execute('SELECT * FROM users WHERE username = %s', (username,))
        user_data = cursor.fetchone()
        
        if user_data:
            print(f"User found in database: {user_data['username']}, role: {user_data['role']}")
            if check_password_hash(user_data['password'], password):
                print("Password verified successfully")
                user = User(user_data)
                login_user(user)
                session['user_id'] = user_data['id']  # Explicitly set session
                print(f"User logged in successfully: {user_data['username']}")
                return redirect(url_for('index'))
            else:
                print("Password verification failed")
        else:
            print(f"User not found: {username}")
        
        cursor.close()
        conn.close()
        
        return render_template('login.html', error='Invalid credentials')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('login'))

# Global variable to track camera status
camera_status = False

def update_attendance_status(cursor, name, current_time, cutoff_time_str):
    """Update a student's attendance status with proper checks"""
    try:
        # Get current time in Vietnam timezone
        cursor.execute("SELECT CONVERT_TZ(NOW(), '+00:00', '+07:00') AS vietnam_time")
        result = cursor.fetchone()
        current_time = result['vietnam_time'] if result else None
        
        if not current_time:
            print("Error: Could not get current time in Vietnam timezone")
            return False
            
        # Get time components
        current_time_str = current_time.strftime('%H:%M:%S')
        time_parts = current_time_str.split(':')
        hours = int(time_parts[0])
        minutes = time_parts[1]
        seconds = time_parts[2]
        
        # Determine status based on time
        status = 'Present' if current_time_str <= cutoff_time_str else 'Late'
        
        # Check if student has already been marked today
        cursor.execute("""
            SELECT DiemDanhStatus, ThoiGianDiemDanh 
            FROM Students 
            WHERE HoVaTen = %s 
            AND DATE(CONVERT_TZ(ThoiGianDiemDanh, '+00:00', '+07:00')) = DATE(%s)
        """, (name, current_time))
        existing_record = cursor.fetchone()
        
        if not existing_record:
            # Update student's attendance
            cursor.execute("""
                UPDATE Students 
                SET DiemDanhStatus = %s, 
                    ThoiGianDiemDanh = %s 
                WHERE HoVaTen = %s
            """, (status, current_time, name))
            
            if cursor.rowcount > 0:
                print(f"Updated attendance for {name}: {status} at {current_time}")
                return True
        else:
            print(f"Student {name} already marked for today at {existing_record['ThoiGianDiemDanh']}")
        return False
        
    except Exception as e:
        print(f"Error updating attendance status: {str(e)}")
        return False

def reset_daily_attendance(cursor):
    """Reset attendance status for a new day"""
    try:
        # Get current date in Vietnam timezone
        cursor.execute("SELECT DATE(CONVERT_TZ(NOW(), '+00:00', '+07:00')) AS today_date")
        today_date = cursor.fetchone()['today_date']
        
        # Reset status for students who haven't been marked today
        cursor.execute("""
            UPDATE Students 
            SET DiemDanhStatus = 'Absent'
            WHERE DATE(CONVERT_TZ(ThoiGianDiemDanh, '+00:00', '+07:00')) < %s
            OR ThoiGianDiemDanh IS NULL
        """, (today_date,))
        
        print(f"Reset attendance status for {cursor.rowcount} students for date {today_date}")
        return True
        
    except Exception as e:
        print(f"Error resetting daily attendance: {str(e)}")
        return False

@app.route('/video_feed')
@login_required
def video_feed():
    global camera_status
    
    def generate():
        global camera_status
        cap = cv2.VideoCapture(0)
        camera_status = True
        print("\n=== Camera Started ===")
        
        # Set basic camera properties
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        
        knn_clf = load_face_recognition_model()
        if knn_clf is None:
            print("No face recognition model found. Please add students first.")
        else:
            print("Face recognition model loaded successfully.")
        
        frame_count = 0
        last_print_time = time.time()
        print_interval = 60  # seconds between prints
        last_reset_time = None
        
        try:
            while True:
                ret, frame = cap.read()
                if not ret:
                    print("Error: Could not read frame from camera")
                    break
                
                # Process every 3rd frame to reduce CPU load
                frame_count += 1
                if frame_count % 3 != 0:
                    continue
                
                # Create a copy of the frame for drawing
                display_frame = frame.copy()
                
                # Resize frame to 1/4 size for faster processing
                small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
                
                # Convert frame to RGB for face recognition
                rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)
                
                # Find faces in frame using HOG model (faster than CNN)
                face_locations = face_recognition.face_locations(rgb_small_frame, model="hog")
                
                # Get current time as float
                current_time = time.time()
                
                # Check if we should print status
                time_since_last_print = current_time - last_print_time
                should_print = time_since_last_print >= print_interval
                
                # Check if we should reset attendance for a new day
                if last_reset_time is None or (current_time - last_reset_time) >= 86400:  # 24 hours
                    conn = get_db_connection()
                    if conn:
                        cursor = conn.cursor(dictionary=True)
                        reset_daily_attendance(cursor)
                        conn.commit()
                        cursor.close()
                        conn.close()
                    last_reset_time = current_time
                
                if face_locations:
                    if should_print:
                        print(f"\n=== Frame {frame_count} at {datetime.now().strftime('%H:%M:%S')} ===")
                        print(f"Found {len(face_locations)} face(s) in frame")
                        last_print_time = current_time
                    
                    # First, draw rectangles for all detected faces
                    for (top, right, bottom, left) in face_locations:
                        # Scale back up face locations since the frame we detected in was scaled to 1/4 size
                        top *= 4
                        right *= 4
                        bottom *= 4
                        left *= 4
                        
                        # Draw a default rectangle for all detected faces
                        cv2.rectangle(display_frame, (left, top), (right, bottom), (128, 128, 128), 2)
                    
                    # Now try to recognize faces if we have a model
                    if knn_clf is not None:
                        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
                        
                        # Process each face
                        for i, face_encoding in enumerate(face_encodings):
                            # Get the face location for this encoding
                            (top, right, bottom, left) = face_locations[i]
                            
                            # Scale back up face locations
                            top *= 4
                            right *= 4
                            bottom *= 4
                            left *= 4
                            
                            # Find closest match
                            closest_distances = knn_clf.kneighbors([face_encoding], n_neighbors=1)
                            distance = closest_distances[0][0][0]
                            is_match = distance <= 0.7
                            
                            if is_match:
                                name = str(knn_clf.predict([face_encoding])[0])
                                
                                if should_print:
                                    print(f"\nFace #{i+1}:")
                                    print(f"  - Recognized as: {name}")
                                    print(f"  - Confidence: {1 - distance:.2%}")
                                
                                # Update attendance status in database
                                try:
                                    conn = get_db_connection()
                                    if conn:
                                        cursor = conn.cursor(dictionary=True)
                                        
                                        # Get current time in Vietnam timezone
                                        cursor.execute("SELECT CONVERT_TZ(NOW(), '+00:00', '+07:00') AS vietnam_time")
                                        result = cursor.fetchone()
                                        current_time = result['vietnam_time'] if result else None
                                        
                                        if current_time:
                                            # Get cutoff time
                                            cursor.execute("SELECT config_value FROM config WHERE config_key = 'cutoff_time'")
                                            cutoff_result = cursor.fetchone()
                                            cutoff_time = "08:30"  # Default cutoff time
                                            if cutoff_result:
                                                cutoff_value = cutoff_result['config_value']
                                                parts = cutoff_value.split(' ')
                                                if len(parts) == 2:
                                                    cutoff_time = parts[1]
                                            
                                            # Convert times to comparable format
                                            current_time_str = current_time.strftime('%H:%M:%S')
                                            cutoff_time_str = cutoff_time + ':00'
                                            
                                            # Update attendance status
                                            if update_attendance_status(cursor, name, current_time, cutoff_time_str):
                                                conn.commit()
                                            
                                        cursor.close()
                                        conn.close()
                                except Exception as e:
                                    print(f"Error updating attendance: {str(e)}")
                                
                                # Update rectangle and add name for recognized face
                                cv2.rectangle(display_frame, (left, top), (right, bottom), (0, 255, 0), 2)
                                cv2.putText(display_frame, name, (left, top - 10), 
                                          cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)
                            else:
                                if should_print:
                                    print(f"\nFace #{i+1}:")
                                    print(f"  - Unknown face")
                                    print(f"  - Closest match distance: {distance:.4f}")
                                
                                # Update rectangle and add "Unknown" for unrecognized face
                                cv2.rectangle(display_frame, (left, top), (right, bottom), (0, 0, 255), 2)
                                cv2.putText(display_frame, "Unknown", (left, top - 10), 
                                          cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
                
                if should_print:
                    last_print_time = current_time
                
                # Convert frame to JPEG
                ret, buffer = cv2.imencode('.jpg', display_frame)
                if not ret:
                    print("Error: Could not encode frame")
                    continue
                
                frame = buffer.tobytes()
                yield (b'--frame\r\n'
                       b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
                
        except Exception as e:
            print(f"\n=== Camera Error: {str(e)} ===")
        finally:
            cap.release()
            camera_status = False
            print("Camera released")
    
    return Response(generate(),
                    mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/attendance')
@login_required
def attendance():
    try:
        conn = get_db_connection()
        if not conn:
            return render_template('attendance.html', 
                                error="Database connection failed", 
                                attendance_data=None, 
                                selected_date=None,
                                dates_in_month=None,
                                stats={},
                                monthly_summary={},
                                cutoff_time=None,
                                gmt=None)

        # Get cutoff time from config
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT config_value FROM config WHERE config_key = 'cutoff_time'")
        cutoff_result = cursor.fetchone()
        
        # Default cutoff time if not set
        cutoff_time = "08:30"
        gmt = "GMT+7"
        
        if cutoff_result:
            cutoff_value = cutoff_result['config_value']
            parts = cutoff_value.split(' ')
            if len(parts) == 2:
                gmt = parts[0]
                cutoff_time = parts[1]

        # Get selected date from query parameters, default to today
        selected_date = request.args.get('date')
        if not selected_date:
            # Use current date
            selected_date = datetime.now().strftime('%Y-%m-%d')
            
        # Get today's date for highlighting
        today_date = datetime.now().strftime('%Y-%m-%d')
        
        # Get all dates in the current month
        current_date = datetime.strptime(selected_date, '%Y-%m-%d')
        year = current_date.year
        month = current_date.month
        
        # Get first and last day of the month
        first_day = datetime(year, month, 1)
        if month == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1) - timedelta(days=1)
        
        # Get all dates in the month
        dates_in_month = []
        current = first_day
        while current <= last_day:
            dates_in_month.append(current.strftime('%Y-%m-%d'))
            current += timedelta(days=1)
        
        # Get attendance data for the selected date
        cursor.execute("""
            SELECT 
                s.UID,
                s.HoVaTen,
                s.Lop,
                s.Gender,
                CASE 
                    WHEN DATE(CONVERT_TZ(s.ThoiGianDiemDanh, '+00:00', '+07:00')) = %s THEN s.DiemDanhStatus
                    ELSE 'Absent'
                END as DiemDanhStatus,
                CASE 
                    WHEN DATE(CONVERT_TZ(s.ThoiGianDiemDanh, '+00:00', '+07:00')) = %s THEN s.ThoiGianDiemDanh
                    ELSE NULL
                END as DiemDanhTime
            FROM 
                students s
            WHERE 
                s.NgaySinh <= %s
            ORDER BY 
                s.HoVaTen
        """, (selected_date, selected_date, selected_date))
        
        attendance_data = cursor.fetchall()
        
        # Initialize stats with default values
        stats = {
            'Present': 0,
            'Late': 0,
            'Absent': 0
        }
        
        # Count attendance statuses
        for record in attendance_data:
            if record['DiemDanhStatus'] in stats:
                stats[record['DiemDanhStatus']] += 1
        
        # Get monthly attendance summary
        cursor.execute("""
            SELECT 
                DATE(CONVERT_TZ(ThoiGianDiemDanh, '+00:00', '+07:00')) as date,
                CASE 
                    WHEN DiemDanhStatus IS NULL THEN 'Absent'
                    ELSE DiemDanhStatus
                END as DiemDanhStatus,
                COUNT(*) as count
            FROM 
                students
            WHERE 
                DATE(CONVERT_TZ(ThoiGianDiemDanh, '+00:00', '+07:00')) BETWEEN %s AND %s
            GROUP BY 
                DATE(CONVERT_TZ(ThoiGianDiemDanh, '+00:00', '+07:00')), 
                CASE 
                    WHEN DiemDanhStatus IS NULL THEN 'Absent'
                    ELSE DiemDanhStatus
                END
        """, (first_day.strftime('%Y-%m-%d'), last_day.strftime('%Y-%m-%d')))
        
        monthly_stats = cursor.fetchall()
        monthly_summary = {}
        for stat in monthly_stats:
            date = stat['date'].strftime('%Y-%m-%d')
            if date not in monthly_summary:
                monthly_summary[date] = {}
            monthly_summary[date][stat['DiemDanhStatus']] = stat['count']
        
        cursor.close()
        conn.close()
        
        return render_template('attendance.html', 
                             attendance_data=attendance_data,
                             selected_date=selected_date,
                             dates_in_month=dates_in_month,
                             stats=stats,
                             monthly_summary=monthly_summary,
                             cutoff_time=cutoff_time,
                             gmt=gmt,
                             today_date=today_date)
                             
    except Exception as e:
        print(f"Error in attendance route: {str(e)}")
        return render_template('attendance.html', 
                             error=f"Error loading attendance data: {str(e)}",
                             attendance_data=None,
                             selected_date=None,
                             dates_in_month=None,
                             stats={},
                             monthly_summary={},
                             cutoff_time=None,
                             gmt=None)

@app.route('/api/attendance')
@login_required
def get_attendance_data():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    try:
        print("Starting attendance data fetch...")
        
        # Get cutoff time from config
        cursor.execute("SELECT config_value FROM config WHERE config_key = 'cutoff_time'")
        cutoff_result = cursor.fetchone()
        print(f"Cutoff result: {cutoff_result}")
        
        # Default cutoff time if not set
        cutoff_time = "08:30"
        gmt = "GMT+7"
        
        if cutoff_result:
            cutoff_value = cutoff_result['config_value']
            parts = cutoff_value.split(' ')
            if len(parts) == 2:
                gmt = parts[0]
                cutoff_time = parts[1]
        print(f"Using cutoff time: {cutoff_time}, GMT: {gmt}")
        
        # Get current date in Vietnam timezone
        cursor.execute("SELECT DATE(CONVERT_TZ(NOW(), '+00:00', '+07:00')) AS today_date")
        today_date_result = cursor.fetchone()
        print(f"Today's date result: {today_date_result}")
        
        if not today_date_result:
            raise Exception("Failed to get today's date")
            
        today_date = today_date_result['today_date']
        print(f"Today's date: {today_date}")
        
        # Get all students with their attendance data
        cursor.execute('''
            SELECT 
                HoVaTen, 
                DiemDanhStatus, 
                DATE_FORMAT(CONVERT_TZ(ThoiGianDiemDanh, '+00:00', '+07:00'), '%Y-%m-%d %H:%i:%s') as ThoiGianDiemDanh
            FROM students 
            ORDER BY 
                CASE 
                    WHEN ThoiGianDiemDanh IS NOT NULL THEN 0 
                    ELSE 1 
                END,
                ThoiGianDiemDanh DESC
        ''')
        attendance_data = cursor.fetchall()
        print(f"Fetched {len(attendance_data)} students")
        
        # Process attendance data to add late status
        for record in attendance_data:
            if record['DiemDanhStatus'] == '✓' and record['ThoiGianDiemDanh']:
                # Convert attendance time to Vietnam timezone
                cursor.execute("""
                    SELECT TIME(CONVERT_TZ(%s, '+00:00', '+07:00')) AS local_time
                """, (record['ThoiGianDiemDanh'],))
                local_time_result = cursor.fetchone()
                print(f"Local time result for {record['HoVaTen']}: {local_time_result}")
                
                if local_time_result and local_time_result['local_time']:
                    # Convert timedelta to string in HH:MM:SS format
                    total_seconds = local_time_result['local_time'].total_seconds()
                    hours = int(total_seconds // 3600)
                    minutes = int((total_seconds % 3600) // 60)
                    seconds = int(total_seconds % 60)
                    local_time = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                    
                    # Check if attendance is late
                    if local_time > cutoff_time:
                        record['DiemDanhStatus'] = 'Late'
                        print(f"Marked {record['HoVaTen']} as late")
        
        response_data = {
            "success": True,
            "attendance_data": attendance_data,
            "cutoff_time": cutoff_time,
            "gmt": gmt,
            "today_date": str(today_date)
        }
        
        print("Sending response:", response_data)
        
        return jsonify(response_data)
                            
    except Exception as e:
        print(f"Error in attendance API: {str(e)}")
        import traceback
        print("Traceback:", traceback.format_exc())
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500
        
    finally:
        cursor.close()
        conn.close()

@app.route('/students')
@login_required
def students():
    try:
        print("Attempting to connect to database...")
        conn = get_db_connection()
        if not conn:
            print("Failed to connect to database")
            return render_template('students.html', students=[], error="Database connection failed")
            
        print("Database connection successful")
        cursor = conn.cursor(dictionary=True)
        
        # First, let's check if the Students table exists
        cursor.execute("SHOW TABLES LIKE 'Students'")
        if not cursor.fetchone():
            print("Students table does not exist")
            return render_template('students.html', students=[], error="Students table not found")
            
        # Now let's check the table structure
        cursor.execute("DESCRIBE Students")
        columns = cursor.fetchall()
        print("Table structure:", columns)
        
        # Let's see what's in the table
        cursor.execute('SELECT * FROM Students ORDER BY UID DESC')
        students = cursor.fetchall()
        print(f"Found {len(students)} students in database")
        for student in students:
            print(f"Student: {student}")
        
        cursor.close()
        conn.close()
        
        return render_template('students.html', students=students)
        
    except Exception as e:
        print(f"Error fetching students: {str(e)}")
        return render_template('students.html', students=[], error=str(e))

@app.route('/api/add_student', methods=['POST'])
@login_required
def add_student():
    conn = None
    cursor = None
    temp_image_path = None
    final_image_path = None
    
    try:
        # Get form data
        full_name = request.form.get('full_name')
        class_name = request.form.get('class')
        gender = request.form.get('gender')
        student_image = request.files.get('photo')
        
        print(f"Received data - Name: {full_name}, Class: {class_name}, Gender: {gender}")
        
        if not all([full_name, class_name, gender, student_image]):
            return jsonify({'success': False, 'message': 'All fields are required'}), 400
            
        # Check for duplicate name
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
            
        cursor = conn.cursor()
        
        # Check if Students table exists, create if not
        cursor.execute("SHOW TABLES LIKE 'Students'")
        if not cursor.fetchone():
            print("Students table not found, creating...")
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS Students (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    UID VARCHAR(50) NOT NULL,
                    HoVaTen VARCHAR(255) NOT NULL,
                    NgaySinh DATE NOT NULL,
                    Lop VARCHAR(50) NOT NULL,
                    Gender ENUM('Male', 'Female') NOT NULL,
                    DiemDanhStatus ENUM('Absent', 'Present', 'Late') DEFAULT 'Absent',
                    ThoiGianDiemDanh DATETIME NULL,
                    ImagePath VARCHAR(255) NOT NULL
                )
            """)
            conn.commit()
            print("Students table created successfully")
        
        cursor.execute('SELECT HoVaTen FROM Students WHERE HoVaTen = %s', (full_name,))
        existing_student = cursor.fetchone()
        if existing_student:
            return jsonify({'success': False, 'message': 'A student with this name already exists'}), 400

        # Generate unique ID for the student
        student_id = f"STU{datetime.now().strftime('%Y%m%d%H%M%S')}"
        print(f"Generated student ID: {student_id}")
        
        # Save the image temporarily to check for duplicate faces
        if student_image:
            try:
                # Create student_images directory if it doesn't exist
                if not os.path.exists('student_images'):
                    os.makedirs('student_images')
                    print("Created student_images directory")
                
                # Save image with temporary filename
                temp_image_path = os.path.join('student_images', f"temp_{student_id}.jpg")
                student_image.save(temp_image_path)
                print(f"Saved temporary image to: {temp_image_path}")
                
                # Load the new image and get face encoding
                new_image = face_recognition.load_image_file(temp_image_path)
                new_face_encodings = face_recognition.face_encodings(new_image)
                
                if not new_face_encodings:
                    os.remove(temp_image_path)
                    return jsonify({'success': False, 'message': 'No face detected in the uploaded image. Please upload a clear photo with a visible face.'}), 400
                
                print("Successfully detected face in new image")
                new_face_encoding = new_face_encodings[0]
                
                # Check for duplicate faces in existing student images
                cursor.execute('SELECT ImagePath, HoVaTen FROM Students')
                existing_images = cursor.fetchall()
                print(f"Found {len(existing_images)} existing student images to check")
                
                for (image_path, student_name) in existing_images:
                    if os.path.exists(image_path):
                        try:
                            existing_image = face_recognition.load_image_file(image_path)
                            existing_face_encodings = face_recognition.face_encodings(existing_image)
                            
                            if existing_face_encodings:
                                # Compare face encodings with a more lenient threshold
                                face_distance = face_recognition.face_distance([existing_face_encodings[0]], new_face_encoding)[0]
                                print(f"Face distance with {student_name}: {face_distance}")
                                if face_distance < 0.4:  # Increased threshold for considering faces as the same
                                    os.remove(temp_image_path)
                                    return jsonify({
                                        'success': False, 
                                        'message': f'This photo appears to be of an existing student: {student_name}. Please upload a different photo.'
                                    }), 400
                        except Exception as e:
                            print(f"Error processing existing image {image_path}: {str(e)}")
                            continue
                
                # If no duplicates found, move the temporary file to its final location
                final_image_path = os.path.join('student_images', f"{student_id}.jpg")
                os.rename(temp_image_path, final_image_path)
                print(f"Moved image to final location: {final_image_path}")
                
            except Exception as e:
                print(f"Error processing image: {str(e)}")
                if temp_image_path and os.path.exists(temp_image_path):
                    os.remove(temp_image_path)
                return jsonify({'success': False, 'message': f'Error processing image: {str(e)}'}), 500
        
        try:
            # Add student to database with initial attendance status
            print(f"Inserting student into database: {student_id}, {full_name}, {class_name}, {gender}, {final_image_path}")
            cursor.execute('''
                INSERT INTO Students (UID, HoVaTen, NgaySinh, Lop, Gender, ImagePath, DiemDanhStatus, ThoiGianDiemDanh)
                VALUES (%s, %s, %s, %s, %s, %s, 'Absent', NULL)
            ''', (student_id, full_name, datetime.now().date(), class_name, gender, final_image_path))
            
            # Verify the insertion
            cursor.execute('SELECT * FROM Students WHERE UID = %s', (student_id,))
            inserted_student = cursor.fetchone()
            if not inserted_student:
                raise Exception("Student was not inserted successfully")
            
            print("Successfully added student to database")
            print(f"Inserted student data: {inserted_student}")
            
            # Commit the changes
            conn.commit()
            print("Database changes committed successfully")
            
        except Exception as e:
            print(f"Database error: {str(e)}")
            if conn:
                conn.rollback()
            if final_image_path and os.path.exists(final_image_path):
                os.remove(final_image_path)
            return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500
        
        try:
            # Retrain the model with the new student using our train_model.py script
            import train_model
            train_result = train_model.train_from_db()
            
            if not train_result:
                return jsonify({'success': False, 'message': 'Student added but model training failed'}), 500
            
            print("Successfully retrained the model")
            return jsonify({'success': True, 'message': 'Student added successfully and model retrained'})
            
        except Exception as e:
            print(f"Model training error: {str(e)}")
            return jsonify({'success': False, 'message': f'Model training error: {str(e)}'}), 500
        
    except Exception as e:
        print(f"Unexpected error in add_student: {str(e)}")
        # Clean up temporary file if it exists
        if temp_image_path and os.path.exists(temp_image_path):
            os.remove(temp_image_path)
        if final_image_path and os.path.exists(final_image_path):
            os.remove(final_image_path)
        return jsonify({'success': False, 'message': f'Unexpected error: {str(e)}'}), 500
        
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/clear_students', methods=['POST'])
def clear_students():
    try:
        # Delete all records from Students table
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM Students')
        conn.commit()
        cursor.close()
        conn.close()
        print("Cleared all student records from database")

        # Delete the model file if it exists
        model_path = "trained_knn_model.clf"
        if os.path.exists(model_path):
            os.remove(model_path)
            print("Deleted old model file")

        # Clear the student_images directory
        student_images_dir = "student_images"
        if os.path.exists(student_images_dir):
            for filename in os.listdir(student_images_dir):
                file_path = os.path.join(student_images_dir, filename)
                try:
                    if os.path.isfile(file_path):
                        os.unlink(file_path)
                except Exception as e:
                    print(f"Error deleting {file_path}: {e}")
            print("Cleared student images directory")

        # Train a new empty model using our train_model.py script
        import train_model
        train_model.train_from_db()
        print("Created new empty model")

        return jsonify({"success": True, "message": "All students cleared successfully"})
    except Exception as e:
        print(f"Error clearing students: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/get_cutoff', methods=['GET'])
@login_required
def get_cutoff():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Create config table if it doesn't exist
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS config (
                config_key VARCHAR(50) PRIMARY KEY,
                config_value VARCHAR(255) NOT NULL
            )
        """)
        
        # Get current cutoff time
        cursor.execute("SELECT config_value FROM config WHERE config_key = 'cutoff_time'")
        result = cursor.fetchone()
        
        if result:
            cutoff_value = result['config_value']
            # Parse the value (format: "GMT+7 08:30")
            parts = cutoff_value.split(' ')
            if len(parts) == 2:
                gmt = parts[0]
                cutoff_time = parts[1]
                return jsonify({
                    "success": True,
                    "gmt": gmt,
                    "cutoff_time": cutoff_time
                })
        
        # Return default values if no cutoff time is set
        return jsonify({
            "success": True,
            "gmt": "GMT+7",
            "cutoff_time": "08:30"
        })
        
    except Exception as e:
        print(f"Error getting cutoff time: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/check_cutoff')
@login_required
def check_cutoff():
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        
        # Get current time in Vietnam timezone
        cursor.execute("""
            SELECT 
                DATE_FORMAT(CONVERT_TZ(NOW(), '+00:00', '+07:00'), '%H:%i:%s') AS vietnam_time,
                DATE_FORMAT(CONVERT_TZ(NOW(), '+00:00', '+07:00'), '%Y-%m-%d') AS vietnam_date
        """)
        current_time_result = cursor.fetchone()
        
        # Hard code the time by subtracting 7 hours
        if current_time_result:
            time_parts = current_time_result['vietnam_time'].split(':')
            hours = int(time_parts[0])
            minutes = time_parts[1]
            seconds = time_parts[2]
            
            # Subtract 7 hours and handle day rollover
            hours = (hours - 7) % 24
            
            current_time = f"{hours:02d}:{minutes}:{seconds}"
            current_date = current_time_result['vietnam_date']
        else:
            current_time = "Error getting current time"
            current_date = "Error getting current date"
        
        # Get current cutoff time
        cursor.execute("SELECT config_value FROM config WHERE config_key = 'cutoff_time'")
        result = cursor.fetchone()
        
        if result:
            cutoff_value = result['config_value']
            # Parse the value (format: "GMT+7 08:30")
            parts = cutoff_value.split(' ')
            if len(parts) == 2:
                gmt = parts[0]
                cutoff_time = parts[1]
                return f"Current date: {current_date}<br>Current time: {current_time} (GMT+7)<br>Current cutoff time: {cutoff_time} ({gmt})"
        
        return f"Current date: {current_date}<br>Current time: {current_time} (GMT+7)<br>No cutoff time set. Using default: 08:30 (GMT+7)"
        
    except Exception as e:
        return f"Error checking times: {str(e)}"
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/set_cutoff', methods=['POST'])
@login_required
def set_cutoff():
    try:
        data = request.get_json()
        gmt = data.get('gmt')
        cutoff_time = data.get('cutoff_time')
        
        if not gmt or not cutoff_time:
            return jsonify({
                "success": False,
                "message": "GMT and cutoff time are required"
            }), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Create config table if it doesn't exist
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS config (
                config_key VARCHAR(50) PRIMARY KEY,
                config_value VARCHAR(255) NOT NULL
            )
        """)
        
        # Update or insert cutoff time
        value = f"{gmt} {cutoff_time}"
        cursor.execute("""
            INSERT INTO config (config_key, config_value)
            VALUES ('cutoff_time', %s)
            ON DUPLICATE KEY UPDATE config_value = %s
        """, (value, value))
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": "Cutoff time updated successfully"
        })
        
    except Exception as e:
        print(f"Error setting cutoff time: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/api/status')
@login_required
def get_status():
    try:
        # Check database connection
        conn = get_db_connection()
        db_status = conn is not None
        if conn:
            conn.close()
            
        # Check face recognition model
        model = load_face_recognition_model()
        model_status = model is not None
        
        return jsonify({
            "success": True,
            "database": {
                "status": db_status,
                "message": "Connected" if db_status else "Disconnected"
            },
            "face_recognition": {
                "status": model_status,
                "message": "Model loaded" if model_status else "No model found"
            }
        })
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500

@app.route('/api/camera_status')
@login_required
def get_camera_status():
    global camera_status
    return jsonify({
        "success": True,
        "camera": {
            "status": camera_status,
            "message": "Camera is active" if camera_status else "Camera is inactive"
        }
    })

@app.route('/camera')
@login_required
def camera_page():
    return render_template('camera.html')

@app.route('/api/stop_camera', methods=['POST'])
@login_required
def stop_camera():
    global camera_status
    camera_status = False
    return jsonify({
        "success": True,
        "message": "Camera stopped successfully"
    })

@app.route('/api/delete_student', methods=['POST'])
@login_required
def delete_student():
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        
        if not student_id:
            return jsonify({
                "success": False,
                "message": "Student ID is required"
            }), 400
            
        conn = get_db_connection()
        if not conn:
            return jsonify({
                "success": False,
                "message": "Database connection failed"
            }), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # First get the student's image path
        cursor.execute("SELECT ImagePath FROM students WHERE UID = %s", (student_id,))
        student = cursor.fetchone()
        
        if not student:
            return jsonify({
                "success": False,
                "message": "Student not found"
            }), 404
            
        # Delete the student from database
        cursor.execute("DELETE FROM students WHERE UID = %s", (student_id,))
        conn.commit()
        
        # Delete the student's image file if it exists
        if student['ImagePath'] and os.path.exists(student['ImagePath']):
            try:
                os.remove(student['ImagePath'])
            except Exception as e:
                print(f"Warning: Could not delete image file: {str(e)}")
        
        # Retrain the model
        try:
            import train_model
            train_model.train_from_db()
        except Exception as e:
            print(f"Warning: Could not retrain model: {str(e)}")
        
        return jsonify({
            "success": True,
            "message": "Student deleted successfully"
        })
        
    except Exception as e:
        print(f"Error deleting student: {str(e)}")
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500
        
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

@app.route('/export_attendance/<date>')
@login_required
def export_attendance(date):
    if current_user.role not in ['admin', 'superuser']:
        return jsonify({
            "success": False,
            "message": "You do not have permission to export attendance data."
        }), 403
    
    try:
        # Get the year and month from the date
        year, month = date.split('-')[:2]
        year = int(year)
        month = int(month)
        
        # Connect to database
        conn = get_db_connection()
        if not conn:
            return jsonify({
                "success": False,
                "message": "Database connection error"
            }), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Get all days in the month
        if month == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1) - timedelta(days=1)
        
        # Create a list of all dates in the month
        dates = []
        current = datetime(year, month, 1)
        while current <= last_day:
            dates.append(current)
            current += timedelta(days=1)
        
        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance {year}-{month}"
        
        # Add headers
        headers = ['ID', 'Name', 'Department', 'Gender']
        for date in dates:
            headers.append(date.strftime('%Y-%m-%d'))
        headers.append('Total Fine')  # Add Fine column header
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Get all students
        query = "SELECT * FROM students ORDER BY UID"
        cursor.execute(query)
        students = cursor.fetchall()
        
        # For each student, add their attendance data
        for student in students:
            row = [
                student['UID'],
                student['HoVaTen'],
                student['Lop'],
                student['Gender']
            ]
            
            # Initialize fine counters
            late_count = 0
            absent_count = 0
            
            # Get employee's creation date
            employee_creation_date = student['NgaySinh']
            
            # Get attendance for each date
            for date in dates:
                # Check if the date is before employee creation
                if date.date() < employee_creation_date:
                    cell_value = "None"
                else:
                    query = """
                        SELECT DiemDanhStatus, TIME(ThoiGianDiemDanh) as time
                        FROM students
                        WHERE UID = %s AND DATE(CONVERT_TZ(ThoiGianDiemDanh, '+00:00', '+07:00')) = %s
                    """
                    cursor.execute(query, (student['UID'], date.strftime('%Y-%m-%d')))
                    attendance = cursor.fetchone()
                    
                    if attendance and attendance['time']:
                        # Convert the time to total seconds and then to hours, minutes, seconds
                        total_seconds = attendance['time'].total_seconds()
                        hours = int(total_seconds // 3600)
                        minutes = int((total_seconds % 3600) // 60)
                        seconds = int(total_seconds % 60)
                        
                        # Add 10 hours and handle overflow
                        new_hour = (hours + 10) % 24
                        
                        time_str = f"{new_hour:02d}:{minutes:02d}:{seconds:02d}"
                        cell_value = f"{attendance['DiemDanhStatus']} ({time_str})"
                        
                        # Count late days
                        if attendance['DiemDanhStatus'] == 'Late':
                            late_count += 1
                    else:
                        cell_value = "Absent"
                        absent_count += 1
                
                row.append(cell_value)
            
            # Calculate total fine (Late: 50, Absent: 100)
            total_fine = (late_count * 50) + (absent_count * 100)
            row.append(f"{total_fine}")  # Add total fine to the row
            
            ws.append(row)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Create a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        
        # Send the file with proper headers
        response = send_file(
            tmp_path,
            as_attachment=True,
            download_name=f"attendance_{year}_{month}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Add headers to force download
        response.headers['Content-Disposition'] = f'attachment; filename=attendance_{year}_{month}.xlsx'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error exporting attendance data: {str(e)}"
        }), 500
    finally:
        if 'conn' in locals():
            conn.close()
        # Clean up the temporary file after sending
        if 'tmp_path' in locals():
            try:
                os.unlink(tmp_path)
            except:
                pass

@app.route('/api/check_attendance/<employee_id>')
@login_required
def check_attendance(employee_id):
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({
                "success": False,
                "message": "Database connection failed"
            }), 500
            
        cursor = conn.cursor(dictionary=True)
        
        # Get employee's attendance data for the current month
        current_date = datetime.now()
        first_day = datetime(current_date.year, current_date.month, 1)
        if current_date.month == 12:
            last_day = datetime(current_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(current_date.year, current_date.month + 1, 1) - timedelta(days=1)
        
        # Get employee's creation date
        cursor.execute("SELECT NgaySinh FROM students WHERE UID = %s", (employee_id,))
        employee = cursor.fetchone()
        if not employee:
            return jsonify({
                "success": False,
                "message": "Employee not found"
            }), 404
            
        employee_creation_date = employee['NgaySinh']
        
        # Get attendance data
        cursor.execute("""
            SELECT 
                DATE(CONVERT_TZ(ThoiGianDiemDanh, '+00:00', '+07:00')) as date,
                DiemDanhStatus,
                TIME(ThoiGianDiemDanh) as time
            FROM students
            WHERE UID = %s 
            AND DATE(CONVERT_TZ(ThoiGianDiemDanh, '+00:00', '+07:00')) BETWEEN %s AND %s
            ORDER BY date
        """, (employee_id, first_day.strftime('%Y-%m-%d'), last_day.strftime('%Y-%m-%d')))
        
        attendance_data = cursor.fetchall()
        
        # Create a list of all dates in the month
        dates = []
        current = first_day
        while current <= last_day:
            dates.append(current.strftime('%Y-%m-%d'))
            current += timedelta(days=1)
        
        # Format the attendance data
        formatted_data = []
        for date in dates:
            if datetime.strptime(date, '%Y-%m-%d').date() < employee_creation_date:
                status = "None"
                time = None
            else:
                # Find matching attendance record
                record = next((r for r in attendance_data if r['date'].strftime('%Y-%m-%d') == date), None)
                if record:
                    status = record['DiemDanhStatus']
                    if record['time']:
                        total_seconds = record['time'].total_seconds()
                        hours = int(total_seconds // 3600)
                        minutes = int((total_seconds % 3600) // 60)
                        seconds = int(total_seconds % 60)
                        new_hour = (hours + 10) % 24
                        time = f"{new_hour:02d}:{minutes:02d}:{seconds:02d}"
                    else:
                        time = None
                else:
                    status = "Absent"
                    time = None
            
            formatted_data.append({
                "date": date,
                "status": status,
                "time": time
            })
        
        return jsonify({
            "success": True,
            "data": formatted_data
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

def init_db():
    try:
        print("Initializing database...")
        conn = mysql.connector.connect(
            host=db_host,
            user=db_username,
            password=db_password
        )
        cursor = conn.cursor()
        
        # Create database if it doesn't exist
        cursor.execute("CREATE DATABASE IF NOT EXISTS attendance_db")
        cursor.execute("USE attendance_db")
        
        # Create users table if it doesn't exist
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INT AUTO_INCREMENT PRIMARY KEY,
                username VARCHAR(255) NOT NULL UNIQUE,
                password VARCHAR(255) NOT NULL,
                role ENUM('superuser', 'admin', 'moderator', 'user') NOT NULL DEFAULT 'user'
            )
        """)
        
        # Create students table if it doesn't exist
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS students (
                id INT AUTO_INCREMENT PRIMARY KEY,
                UID VARCHAR(50) NOT NULL,
                HoVaTen VARCHAR(255) NOT NULL,
                NgaySinh DATE NOT NULL,
                Lop VARCHAR(50) NOT NULL,
                Gender ENUM('Male', 'Female') NOT NULL,
                DiemDanhStatus ENUM('Absent', 'Present', 'Late') DEFAULT 'Absent',
                ThoiGianDiemDanh DATETIME NULL,
                ImagePath VARCHAR(255) NOT NULL
            )
        """)
        
        # Create default users if none exist
        cursor.execute("SELECT COUNT(*) FROM users")
        user_count = cursor.fetchone()[0]
        print(f"Found {user_count} existing users")
        
        if user_count == 0:
            print("Creating default users...")
            default_users = [
                ("superuser", generate_password_hash("superpass"), "superuser"),
                ("admin", generate_password_hash("adminpass"), "admin"),
                ("moderator", generate_password_hash("modpass"), "moderator"),
                ("user", generate_password_hash("userpass"), "user")
            ]
            cursor.executemany(
                "INSERT INTO users (username, password, role) VALUES (%s, %s, %s)",
                default_users
            )
            print("Default users created successfully")
            
            # Verify admin user was created
            cursor.execute("SELECT * FROM users WHERE username = 'admin'")
            admin_user = cursor.fetchone()
            if admin_user:
                print("Admin user verified in database")
            else:
                print("WARNING: Admin user not found in database after creation")
        
        conn.commit()
        cursor.close()
        conn.close()
        print("Database initialization completed")
        
    except Exception as e:
        print(f"Error initializing database: {str(e)}")
        raise e

# Add custom filter for date calculations
@app.template_filter('date_add')
def date_add(date_str, days):
    try:
        date = datetime.strptime(date_str, '%Y-%m-%d')
        new_date = date + timedelta(days=days)
        return new_date.strftime('%Y-%m-%d')
    except:
        return date_str

if __name__ == '__main__':
    init_db()  # Initialize database before running the app
    
    # Check if model exists, if not and there are students, train the model
    model_path = "trained_knn_model.clf"
    if not os.path.exists(model_path):
        print("Model file not found. Checking if training is needed...")
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM Students')
            student_count = cursor.fetchone()[0]
            cursor.close()
            conn.close()
            
            if student_count > 0:
                print(f"Found {student_count} students but no model. Training model...")
                import train_model
                train_model.train_from_db()
            else:
                print("No students found in database. Model will be trained when students are added.")
    
    app.run(debug=True) 